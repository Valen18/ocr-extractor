"""
OCR Extractor — Aplicacion standalone para extraer datos de facturas con IA.

Uso:
    pip install -r requirements.txt
    cp .env.example .env   # configurar GEMINI_API_KEY
    python app.py
"""

import json
import os
import uuid
from pathlib import Path

import pandas as pd
from dotenv import load_dotenv
from flask import Flask, render_template, request, jsonify, send_file

from ocr_service import (
    OCRService, extract_date_from_filename,
    DEFAULT_COLUMNS, COLUMN_LABELS,
)
from reference_matcher import (
    parse_excel_to_records, save_reference, get_reference_status,
    clear_reference, enrich_facturas,
)

load_dotenv()

app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = 50 * 1024 * 1024  # 50MB

BASE_DIR = Path(__file__).resolve().parent
UPLOAD_DIR = BASE_DIR / "uploads"
OUTPUT_DIR = BASE_DIR / "output"
CONFIG_FILE = BASE_DIR / "column_config.json"
UPLOAD_DIR.mkdir(exist_ok=True)
OUTPUT_DIR.mkdir(exist_ok=True)

DB_URL = os.environ.get("DATABASE_URL", "")


def _get_db():
    import psycopg2
    return psycopg2.connect(DB_URL)


def _ensure_config_table():
    if not DB_URL:
        return
    try:
        conn = _get_db()
        with conn.cursor() as cur:
            cur.execute("""
                CREATE TABLE IF NOT EXISTS app_config (
                    key VARCHAR(100) PRIMARY KEY,
                    value JSONB NOT NULL
                )
            """)
            conn.commit()
        conn.close()
    except Exception:
        pass


_config_table_ready = False


def load_column_config() -> dict | None:
    """Carga la configuracion de columnas guardada."""
    global _config_table_ready
    if DB_URL:
        try:
            if not _config_table_ready:
                _ensure_config_table()
                _config_table_ready = True
            conn = _get_db()
            with conn.cursor() as cur:
                cur.execute("SELECT value FROM app_config WHERE key = 'column_config'")
                row = cur.fetchone()
            conn.close()
            if row:
                return row[0] if isinstance(row[0], dict) else json.loads(row[0])
        except Exception:
            pass

    if CONFIG_FILE.exists():
        try:
            return json.loads(CONFIG_FILE.read_text(encoding="utf-8"))
        except Exception:
            pass
    return None


def save_column_config(columns: list, labels: dict, prompt: str = ""):
    """Guarda la configuracion de columnas."""
    data = {"columns": columns, "column_labels": labels, "custom_prompt": prompt}

    # Always save local copy
    CONFIG_FILE.write_text(
        json.dumps(data, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )

    # Also save to DB if available
    if DB_URL:
        try:
            _ensure_config_table()
            conn = _get_db()
            with conn.cursor() as cur:
                cur.execute("""
                    INSERT INTO app_config (key, value) VALUES ('column_config', %s)
                    ON CONFLICT (key) DO UPDATE SET value = %s
                """, (json.dumps(data), json.dumps(data)))
                conn.commit()
            conn.close()
        except Exception:
            pass

_ocr_service = None


def get_service() -> OCRService:
    global _ocr_service
    if _ocr_service is None:
        _ocr_service = OCRService()
    return _ocr_service


# ── Routes ────────────────────────────────────────────────────────────


@app.route("/")
def index():
    service = get_service()
    saved = load_column_config()
    ref_status = get_reference_status()
    return render_template(
        "index.html",
        ocr_available=service.available,
        default_columns=DEFAULT_COLUMNS,
        column_labels=COLUMN_LABELS,
        saved_config=saved,
        reference_loaded=ref_status["loaded"],
        reference_count=ref_status["count"],
    )


@app.route("/reference/upload", methods=["POST"])
def reference_upload():
    """Sube el Excel de referencia con las reservas propias."""
    if "file" not in request.files:
        return jsonify({"ok": False, "error": "No se envio archivo."}), 400

    file = request.files["file"]
    if not file.filename:
        return jsonify({"ok": False, "error": "Archivo sin nombre."}), 400

    ext = Path(file.filename).suffix.lower()
    if ext not in {".xlsx", ".xls"}:
        return jsonify({"ok": False, "error": f"Formato no soportado: {ext}. Usa Excel (.xlsx o .xls)."}), 400

    # Save file temporarily
    tmp_path = UPLOAD_DIR / f"ref_{uuid.uuid4().hex[:8]}{ext}"
    file.save(str(tmp_path))

    try:
        records = parse_excel_to_records(tmp_path)
        if not records:
            return jsonify({"ok": False, "error": "El Excel no contiene datos validos (nombres/localizadores)."}), 400

        count = save_reference(records)

        return jsonify({
            "ok": True,
            "count": count,
            "filename": file.filename,
            "message": f"{count} reservas cargadas correctamente.",
        })

    except Exception as e:
        return jsonify({"ok": False, "error": f"Error al procesar Excel: {str(e)}"}), 500

    finally:
        try:
            tmp_path.unlink(missing_ok=True)
        except Exception:
            pass


@app.route("/reference/status")
def reference_status():
    """Devuelve el estado de la referencia cargada."""
    return jsonify(get_reference_status())


@app.route("/reference/clear", methods=["POST"])
def reference_clear():
    """Borra la referencia cargada."""
    clear_reference()
    return jsonify({"ok": True})


@app.route("/process", methods=["POST"])
def process_file():
    """Procesa un archivo con OCR Vision."""
    service = get_service()

    if not service.available:
        return jsonify({"ok": False, "error": "Gemini no configurado. Anade GEMINI_API_KEY a .env"}), 400

    if "file" not in request.files:
        return jsonify({"ok": False, "error": "No se envio archivo."}), 400

    file = request.files["file"]
    if not file.filename:
        return jsonify({"ok": False, "error": "Archivo sin nombre."}), 400

    ext = Path(file.filename).suffix.lower()
    if ext not in {".pdf", ".jpg", ".jpeg", ".png"}:
        return jsonify({"ok": False, "error": f"Formato no soportado: {ext}. Usa PDF, JPG o PNG."}), 400

    # Save temp file
    filename = f"ocr_{uuid.uuid4().hex[:8]}{ext}"
    file_path = UPLOAD_DIR / filename
    file.save(str(file_path))

    # Get custom prompt from session config
    custom_prompt = request.form.get("custom_prompt", "")

    try:
        facturas = service.extract(file_path, custom_prompt)

        # Fallback: extract date from filename
        filename_date = extract_date_from_filename(file.filename or "")
        if filename_date:
            for fac in facturas:
                if not fac.fecha:
                    fac.fecha = filename_date

        facturas_data = [f.to_dict() for f in facturas]

        # Enrich with reference (formats names + adds nuestro_localizador)
        facturas_data = enrich_facturas(facturas_data)

        return jsonify({
            "ok": True,
            "filename": file.filename,
            "facturas": facturas_data,
            "num_facturas": len(facturas_data),
        })

    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500

    finally:
        # Cleanup temp file
        try:
            file_path.unlink(missing_ok=True)
        except Exception:
            pass


@app.route("/customize-columns", methods=["POST"])
def customize_columns():
    """Interpreta la peticion del usuario para configurar columnas."""
    service = get_service()
    data = request.get_json() or {}
    user_request = data.get("prompt", "")

    if not user_request:
        return jsonify({
            "columns": DEFAULT_COLUMNS,
            "column_labels": COLUMN_LABELS,
            "prompt_extra": "",
            "explanation": "No se especifico ninguna peticion.",
        })

    result = service.customize_columns(user_request)

    # Persist config
    if result.get("columns"):
        save_column_config(
            result["columns"],
            result.get("column_labels", {}),
            result.get("prompt_extra", ""),
        )

    return jsonify(result)


@app.route("/save-columns", methods=["POST"])
def save_columns():
    """Guarda la configuracion de columnas (cuando el usuario quita/anade manualmente)."""
    data = request.get_json() or {}
    columns = data.get("columns", [])
    labels = data.get("column_labels", {})
    prompt = data.get("custom_prompt", "")

    if not columns:
        # Reset to defaults
        if CONFIG_FILE.exists():
            CONFIG_FILE.unlink()
        return jsonify({"ok": True, "reset": True})

    save_column_config(columns, labels, prompt)
    return jsonify({"ok": True})


@app.route("/export", methods=["POST"])
def export_excel():
    """Exporta los datos de la tabla a Excel."""
    data = request.get_json() or {}
    rows = data.get("facturas", [])
    columns = data.get("columns", [])
    column_labels = data.get("column_labels", COLUMN_LABELS)

    if not rows:
        return jsonify({"ok": False, "error": "No hay datos para exportar."}), 400

    df = pd.DataFrame(rows)

    # Reorder and filter columns if specified
    if columns:
        existing = [c for c in columns if c in df.columns]
        extra = [c for c in df.columns if c not in columns]
        df = df[existing + extra]

    # Rename columns to human-readable labels
    rename_map = {k: v for k, v in column_labels.items() if k in df.columns}
    df = df.rename(columns=rename_map)

    output_path = OUTPUT_DIR / "facturas_extraidas.xlsx"

    # Style the Excel output
    with pd.ExcelWriter(str(output_path), engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Facturas")
        ws = writer.sheets["Facturas"]

        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        header_fill = PatternFill(start_color="1a56db", end_color="1a56db", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        thin_border = Border(
            left=Side(style="thin", color="d1d5db"),
            right=Side(style="thin", color="d1d5db"),
            top=Side(style="thin", color="d1d5db"),
            bottom=Side(style="thin", color="d1d5db"),
        )

        for col_idx, cell in enumerate(ws[1], 1):
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center")

        # Auto-width columns
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    val = str(cell.value) if cell.value else ""
                    max_len = max(max_len, len(val))
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = min(max(max_len + 3, 10), 50)

        ws.auto_filter.ref = ws.dimensions
        ws.freeze_panes = "A2"

    return jsonify({"ok": True, "total": len(rows)})


@app.route("/download")
def download():
    path = OUTPUT_DIR / "facturas_extraidas.xlsx"
    if not path.exists():
        return jsonify({"ok": False, "error": "No hay archivo para descargar."}), 404

    return send_file(
        str(path),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="facturas_extraidas.xlsx",
    )


ADMIN_PASSWORD = "Consum0!@2027!"


@app.route("/admin", methods=["GET", "POST"])
def admin_login():
    """Login del backoffice."""
    if request.method == "POST":
        pwd = request.form.get("password", "")
        if pwd == ADMIN_PASSWORD:
            from usage_tracker import get_daily_summary, get_totals, get_all_entries
            return render_template(
                "admin.html",
                authenticated=True,
                daily=get_daily_summary(),
                totals=get_totals(),
                entries=get_all_entries()[:200],
                password=pwd,
            )
        return render_template("admin.html", authenticated=False, error=True)
    return render_template("admin.html", authenticated=False)


@app.route("/admin/api", methods=["POST"])
def admin_api():
    """API JSON para el dashboard (refrescar sin recargar)."""
    data = request.get_json() or {}
    if data.get("password") != ADMIN_PASSWORD:
        return jsonify({"error": "No autorizado"}), 401

    from usage_tracker import get_daily_summary, get_totals, get_all_entries
    return jsonify({
        "daily": get_daily_summary(),
        "totals": get_totals(),
        "entries": get_all_entries()[:200],
    })


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5001))
    print(f"\n  OCR Extractor corriendo en http://localhost:{port}\n")
    app.run(debug=True, port=port)
